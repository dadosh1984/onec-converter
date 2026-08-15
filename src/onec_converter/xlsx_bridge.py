"""XLSX-мост для переноса между разными конфигурациями (пилот.

Выгрузка таблицы источника в XLSX с человекочитаемыми именами реквизитов
и разрешением ссылок (GUID -> наименование), затем конвертация заполненного
XLSX-шаблона приёмника обратно в intermediate JSON для load.

Поток: export_object_xlsx(источник) -> ручной маппинг в Excel ->
       xlsx_to_intermediate(xlsx приёмника) -> load --direct.

Ограничение: переносятся только пользовательские записи (не предопределённые,
не пустые); ссылки в XLSX хранятся как 'Тип:ключ' (как в intermediate),
чтобы load смог резолвить их в ID приёмника.
"""
from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from typing import Any, cast

from openpyxl.worksheet.worksheet import Worksheet

from .source_8x_file import Database1CD, read_metadata

# служебные поля, не переносимые (как в load_8x.object_to_row)
SKIP_FIELDS = {'_VERSION', '_MARKED', '_ISMETADATA', '_FOLDER', '_ORDERFIELD',
               '_PREDEFINEDID'}


def _field_names(db: Database1CD, table: str) -> dict[str, str]:
    """field -> человекочитаемое имя реквизита (физические _FldNNN остаются)."""
    try:
        md = read_metadata(str(db.path)) if hasattr(db, 'path') else None
    except (OSError, ValueError, UnicodeDecodeError):
        md = None
    if not md:
        return {}
    for o in md.get('objects', []):
        if o.get('table') == table:
            return {a['field']: a.get('name') or a['field']
                    for a in o.get('attributes', [])}
    return {}


def export_object_xlsx(source_dir: str | Path, table: str,
                       out: str | Path, limit: int = 0,
                       resolve_refs: bool = True,
                       obj_type: str = '') -> dict[str, Any]:
    """Выгрузить пользовательские записи таблицы источника в XLSX.

    Заголовки — имена реквизитов (не _FldNNN, где есть); ссылочные поля —
    'Тип:ключ' (GUID источника разрешается в код|наименование), строковые —
    как есть. Предопределённые (_ISMETADATA=1) и пустые записи пропускаются.
    Родитель резолвится в `obj_type:код|наименование` (если obj_type задан),
    остальные ссылки — в 'ref:наименование' для ручного маппинга.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font as XFont

    cd = Path(source_dir) / '1Cv8.1CD'
    out_path = Path(out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with Database1CD(cd) as db:
        t = db.tables.get(table)
        if t is None:
            raise KeyError(f'нет таблицы {table!r}')
        fnames = _field_names(db, table)
        # маппинг: имя реквизита -> (field, fdef)
        cols: list[tuple[str, Any]] = []
        for fname, fdef in t.fields.items():
            if fname in SKIP_FIELDS:
                continue
            label = fnames.get(fname, fname)
            if label == 'ID':
                continue
            cols.append((label, fdef))

        wb = Workbook()
        ws = cast(Worksheet, wb.active)
        ws.title = table[:31]
        ws.append([label for label, _ in cols])
        for c in ws[1]:
            c.font = XFont(bold=True)

        rows = 0
        for row in db.table_rows(t):
            if t.fields.get('_ISMETADATA'):
                rec = _row_rec(row, t)
                if rec.get('_ISMETADATA'):
                    continue
            else:
                rec = _row_rec(row, t)
            # пропуск пустых (нет кода и наименования)
            if not _has_content(rec, t):
                continue
            values = []
            for label, fdef in cols:
                raw = row[fdef.offset:fdef.offset + fdef.size]
                val = _decode(fdef, raw)
                if fdef.type in ('RV', 'B') and val and val != '00000000-0000-0000-0000-000000000000':
                    if label == 'Родитель' and obj_type:
                        name = db.ref_name(table, raw)
                        val = f'{obj_type}:{name}' if name else val
                    elif resolve_refs:
                        name = db.ref_name(table, raw)
                        val = f'ref:{name}' if name else val
                    else:
                        val = _guid_key(db, table, raw)
                values.append(val)
            ws.append(values)
            rows += 1
            if limit and rows >= limit:
                break
        wb.save(out_path)
    return {'ok': True, 'path': str(out_path), 'rows': rows, 'table': table}


def _row_rec(row: bytes, t: Any) -> dict[str, Any]:
    from .source_8x_file import decode_field
    rec: dict[str, Any] = {}
    for fname, fdef in t.fields.items():
        try:
            rec[fname] = decode_field(fdef, row[fdef.offset:fdef.offset + fdef.size])
        except (IndexError, ValueError, UnicodeDecodeError):
            rec[fname] = None
    return rec


def _decode(fdef: Any, raw: bytes) -> Any:
    from .source_8x_file import decode_field
    try:
        return decode_field(fdef, raw)
    except (IndexError, ValueError, UnicodeDecodeError):
        return None


def _has_content(rec: dict[str, Any], t: Any) -> bool:
    code = rec.get('_CODE')
    desc = rec.get('_DESCRIPTION')
    if code not in (None, '', 0) or desc not in (None, ''):
        return True
    # документы: есть дата/номер — это пользовательская запись
    for f in ('_DATE_TIME', '_NUMBER'):
        if rec.get(f) not in (None, '', 0):
            return True
    return False


def _guid_key(db: Database1CD, table: str, raw: bytes) -> str:
    """GUID -> 'код|наименование' целевой записи (для читаемости в Excel)."""
    name = db.ref_name(table, raw)
    return name or raw.hex()


def ref_columns_for(target_dir: str | Path, obj_type: str) -> list[str]:
    """Имена ref-реквизитов объекта приёмника (для xlsx_to_intermediate)."""
    cd = Path(target_dir) / '1Cv8.1CD'
    with Database1CD(cd) as db:
        md = read_metadata(str(db.path))
    kind, _, name = obj_type.partition('.')
    for o in md.get('objects', []):
        if o.get('kind') == kind and o.get('name') == name:
            return [a.get('name') or '' for a in o.get('attributes', [])
                    if a.get('type') in ('ref', 'B')]
    return []


def xlsx_to_intermediate(xlsx_path: str | Path,
                         obj_type: str,
                         ref_columns: Iterable[str] | None = None) -> list[dict[str, Any]]:
    """Конвертация XLSX-шаблона приёмника в intermediate JSON для load.

    Каждая строка -> объект {type, id, key, attributes, references}.
    key: [Код, Наименование] если колонки есть, иначе пустой.
    Ссылочные колонки (переданные ref_columns) интерпретируются как
    references, остальные — атрибуты. Пустые строки пропускаются.
    """
    from openpyxl import load_workbook

    ref_cols = set(ref_columns or [])
    wb = load_workbook(Path(xlsx_path))
    ws = cast(Worksheet, wb.active)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    out: list[dict[str, Any]] = []
    for r in rows[1:]:
        if r is None or all(v in (None, '') for v in r):
            continue
        attrs: dict[str, Any] = {}
        refs: dict[str, str] = {}
        for h, v in zip(headers, r):
            if not h or v in (None, ''):
                continue
            if h in ref_cols:
                refs[h] = str(v)
            else:
                attrs[h] = v
        key = [str(attrs.pop('Код')) if 'Код' in attrs else '',
               str(attrs.pop('Наименование')) if 'Наименование' in attrs else '']
        key = [k for k in key if k]
        out.append({
            'type': obj_type,
            'id': f'xlsx:{len(out)}',
            'key': key,
            'attributes': attrs,
            'references': refs,
        })
    return out
