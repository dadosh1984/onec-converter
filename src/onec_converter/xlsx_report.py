"""Человекочитаемый xlsx-отчёт по выгруженным данным (openpyxl).

Один лист на тип объекта; колонки — реквизиты; строки — записи.
Используется для верификации выборки человеком до загрузки в приёмник.
"""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from .intermediate import OBJ_ATTRS, OBJ_KEY, OBJ_TYPE


def build_report(objects: Iterable[dict[str, Any]], out_path: str | Path,
                 max_rows_per_sheet: int = 100_000) -> Path:
    """Сформировать xlsx-отчёт: лист на тип объекта."""
    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]
    grouped: dict[str, list[dict[str, Any]]] = {}
    for obj in objects:
        grouped.setdefault(obj[OBJ_TYPE], []).append(obj)

    header_font = Font(bold=True)
    for obj_type, items in grouped.items():
        ws = wb.create_sheet(title=_sheet_title(obj_type))
        attr_names: list[str] = []
        for it in items:
            for name in it[OBJ_ATTRS]:
                if name not in attr_names:
                    attr_names.append(name)
        headers = ['Ключ'] + attr_names
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
        for it in items[:max_rows_per_sheet]:
            row = ['|'.join(str(p) for p in it[OBJ_KEY])]
            for name in attr_names:
                row.append(it[OBJ_ATTRS].get(name))
            ws.append(row)
    out = Path(out_path)
    wb.save(out)
    return out


def _sheet_title(obj_type: str) -> str:
    # имена листов до 31 символа, без запрещённых символов
    title = obj_type.replace('.', '_').replace(':', '_')[:31]
    return title or 'Objects'


# --- Фаза 8: отчёты структур и размеров таблиц ------------------------------

def build_structure_report(diff: dict[str, Any], out_path: str | Path) -> Path:
    """XLSX-отчёт структуры: листы «Только в источнике/приёмнике» и «Расхождения типов».

    diff — словарь как в выводе compare_structures/diff_structures:
    {only_source: [str], only_target: [str], type_mismatch: [{object, attr,
    source_type, target_type}]}. Пустые секции — лист с заголовками и 0 строк.
    """
    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]
    header_font = Font(bold=True)

    def add_sheet(title: str, headers: list[str], rows: Iterable[list[Any]]) -> None:
        ws = wb.create_sheet(title=_sheet_title(title))
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
        for row in rows:
            ws.append(row)

    only_source = diff.get('only_source', [])
    only_target = diff.get('only_target', [])
    mismatch = diff.get('type_mismatch', [])

    add_sheet('Только в источнике', ['Объект'], ([k] for k in only_source))
    add_sheet('Только в приёмнике', ['Объект'], ([k] for k in only_target))
    add_sheet('Расхождения типов', ['Объект', 'Поле', 'Тип источника', 'Тип приёмника'],
              ([m['object'], m['attr'], m['source_type'], m['target_type']] for m in mismatch))

    out = Path(out_path)
    wb.save(out)
    return out


def build_sizes_report(sizes: list[tuple[str, int, int]], out_path: str | Path,
                       top_n: int = 50) -> Path:
    """XLSX-отчёт размеров таблиц: лист «Таблицы», сортировка по байтам, топ-N.

    sizes — список (имя_таблицы, строки, байты) из Database1CD.table_stats.
    """
    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]
    header_font = Font(bold=True)
    ws = wb.create_sheet(title='Таблицы')
    ws.append(['Таблица', 'Строки', 'Байты'])
    for cell in ws[1]:
        cell.font = header_font
    for name, rows, size in sorted(sizes, key=lambda x: x[2], reverse=True)[:top_n]:
        ws.append([name, rows, size])
    out = Path(out_path)
    wb.save(out)
    return out
