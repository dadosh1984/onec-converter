"""Формат xlsx-моста переноса данных (аналог макета «МакетСохраненияНастроек» .epf).

bridge.xlsx, листы:
- «Настройки»: шапка (версия, режим, объект, флаги, первая строка данных),
  маппинг колонок C1–C11 (совместим с семантикой epf), подвал (события-хуки);
- «Данные»: R1 — заголовки по НомерКолонки, строки данных.

Потоки: ИБ -> xlsx (export) и xlsx -> ИБ (import через intermediate + load).
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, cast

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font as XFont
from openpyxl.worksheet.worksheet import Worksheet

from .typify import KIND_STRING, TypeSpec, parse_type_desc, type_to_text

MODE_CATALOG = 0    # в справочник
MODE_TABLE = 1      # в табличную часть
MODE_REGISTER = 2   # в регистр сведений

MODE_TEXT = {
    MODE_CATALOG: 'в справочник',
    MODE_TABLE: 'в табличную часть',
    MODE_REGISTER: 'в регистр сведений',
}
MODE_FROM_TEXT = {v: k for k, v in MODE_TEXT.items()}

SETTINGS_SHEET = 'Настройки'
DATA_SHEET = 'Данные'
VERSION = '1.2'

# строки шапки листа «Настройки» (0-индекс = номер строки - 1)
H_ROW_VERSION = 0
H_ROW_MODE = 1
H_ROW_OBJECT = 2
H_ROW_NO_NEW = 3
H_ROW_REPLACE = 4
H_ROW_MANUAL = 5
H_ROW_FIRST_DATA = 6
H_ROW_MAPPING_HEADER = 8
H_ROW_MAPPING_FIRST = 9

MAPPING_HEADERS = ['Пометка', 'ИмяРеквизита', 'ПолеПоиска', 'ОписаниеТипов',
                   'РежимЗагрузки', 'ЗначениеПоУмолчанию', 'ИскатьПо/Выражение',
                   'СвязьПоВладельцу', 'СвязьПоТипу', 'ЭлементСвязиПоТипу',
                   'НомерКолонки']

FOOTER_BEFORE = 'ПередЗаписьюОбъекта'
FOOTER_AFTER = 'ПриЗаписиОбъекта'
FOOTER_AFTER_ROW = 'ПослеДобавленияСтроки'


@dataclass
class ColumnSpec:
    """Одна строка маппинга (C1–C11 макета настроек epf)."""

    flag: bool                # Пометка: колонка участвует в загрузке
    attr: str                 # ИмяРеквизита
    search: bool              # ПолеПоиска: участвует в find-or-create
    type_spec: TypeSpec       # ОписаниеТипов
    mode: str = 'Устанавливать'  # 'Устанавливать' | 'Вычислять'
    default: str = ''         # ЗначениеПоУмолчанию (текст)
    lookup: str = ''          # ИскатьПо (или Выражение при mode='Вычислять')
    owner_ref: str = ''       # СвязьПоВладельцу
    type_ref: str = ''        # СвязьПоТипу
    type_elem: int = 0        # ЭлементСвязиПоТипу
    col_num: int = 0          # НомерКолонки (0 = нет колонки)


@dataclass
class BridgeConfig:
    """Полная конфигурация моста (лист «Настройки»)."""

    version: str = VERSION
    mode: int = MODE_CATALOG
    obj_fullname: str = ''    # 'Справочник.Контрагенты'
    no_new: bool = False      # НеСоздаватьНовыхЭлементов
    replace: bool = False     # ЗамещатьСуществующиеЗаписи
    manual_cols: bool = False  # РучнаяНумерацияКолонок
    first_data_row: int = 2   # ПерваяСтрокаДанных (в листе «Данные»)
    columns: list[ColumnSpec] = field(default_factory=list)
    before_write: str = ''    # событие-хук (Python)
    after_write: str = ''
    after_add_row: str = ''


def write_bridge(path: str | Path, cfg: BridgeConfig,
                 data_rows: list[list[Any]]) -> None:
    """Запись xlsx-моста: лист «Настройки» + лист «Данные»."""
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = SETTINGS_SHEET
    ws.append([cfg.version])
    ws.append([MODE_TEXT.get(cfg.mode, MODE_TEXT[MODE_CATALOG])])
    ws.append([cfg.obj_fullname])
    ws.append([_flag_text(cfg.no_new)])
    ws.append([_flag_text(cfg.replace)])
    ws.append([_flag_text(cfg.manual_cols)])
    ws.append([cfg.first_data_row])
    ws.append([])
    ws.append(MAPPING_HEADERS)
    for cell in ws[H_ROW_MAPPING_HEADER + 1]:
        cell.font = XFont(bold=True)
    for col in cfg.columns:
        ws.append([
            _flag_text(col.flag),
            col.attr,
            _flag_text(col.search),
            type_to_text(col.type_spec),
            col.mode,
            col.default,
            col.lookup,
            col.owner_ref,
            col.type_ref,
            col.type_elem or '',
            col.col_num or '',
        ])
    ws.append([])
    if cfg.before_write:
        ws.append([FOOTER_BEFORE, cfg.before_write])
    if cfg.after_write:
        ws.append([FOOTER_AFTER, cfg.after_write])
    if cfg.after_add_row:
        ws.append([FOOTER_AFTER_ROW, cfg.after_add_row])

    dsheet = cast(Worksheet, wb.create_sheet(DATA_SHEET))
    nums = [c.col_num for c in cfg.columns]
    if any(nums):
        # заголовки и данные выравниваются по НомерКолонки (как в epf)
        width = max(nums)
        header: list[Any] = [None] * width
        for c in cfg.columns:
            if c.col_num:
                header[c.col_num - 1] = c.attr
        dsheet.append(header)
        for row in data_rows:
            out_row: list[Any] = [None] * width
            for c, v in zip(cfg.columns, row):
                if c.col_num:
                    out_row[c.col_num - 1] = v
            dsheet.append(out_row)
    else:
        dsheet.append([c.attr for c in cfg.columns])
        for row in data_rows:
            dsheet.append(list(row))

    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def read_bridge(path: str | Path) -> tuple[BridgeConfig, list[list[Any]]]:
    """(конфиг, строки данных) из xlsx-моста."""
    wb = load_workbook(Path(path))
    cfg = _read_settings(cast(Worksheet, wb[SETTINGS_SHEET]))
    data = _read_data(cast(Worksheet, wb[DATA_SHEET]), cfg.first_data_row)
    return cfg, data


def _read_settings(ws: Worksheet) -> BridgeConfig:
    rows = list(ws.iter_rows(values_only=True))
    cfg = BridgeConfig()
    cfg.version = str(rows[H_ROW_VERSION][0] or '') if rows else VERSION
    if len(rows) > H_ROW_MODE:
        cfg.mode = MODE_FROM_TEXT.get(str(rows[H_ROW_MODE][0] or '').strip(),
                                      MODE_CATALOG)
    if len(rows) > H_ROW_OBJECT:
        cfg.obj_fullname = str(rows[H_ROW_OBJECT][0] or '')
    if len(rows) > H_ROW_NO_NEW:
        cfg.no_new = _flag(rows[H_ROW_NO_NEW][0])
    if len(rows) > H_ROW_REPLACE:
        cfg.replace = _flag(rows[H_ROW_REPLACE][0])
    if len(rows) > H_ROW_MANUAL:
        cfg.manual_cols = _flag(rows[H_ROW_MANUAL][0])
    if len(rows) > H_ROW_FIRST_DATA:
        cfg.first_data_row = _int(rows[H_ROW_FIRST_DATA][0]) or 2
    for r in rows[H_ROW_MAPPING_FIRST:]:
        if r is None:
            break
        name = r[1] if len(r) > 1 else None
        if not name or not str(name).strip():
            break
        cfg.columns.append(ColumnSpec(
            flag=_flag(r[0]),
            attr=str(name).strip(),
            search=_flag(r[2]) if len(r) > 2 else False,
            type_spec=_parse_type(r[3] if len(r) > 3 else None),
            mode=str(r[4]).strip() or 'Устанавливать' if len(r) > 4 else 'Устанавливать',
            default=str(r[5] or '') if len(r) > 5 else '',
            lookup=str(r[6] or '') if len(r) > 6 else '',
            owner_ref=str(r[7] or '') if len(r) > 7 else '',
            type_ref=str(r[8] or '') if len(r) > 8 else '',
            type_elem=_int(r[9] if len(r) > 9 else None),
            col_num=_int(r[10] if len(r) > 10 else None),
        ))
    for r in rows:
        if not r or r[0] is None:
            continue
        key = str(r[0]).strip()
        val = str(r[1] or '') if len(r) > 1 else ''
        if key == FOOTER_BEFORE:
            cfg.before_write = val
        elif key == FOOTER_AFTER:
            cfg.after_write = val
        elif key == FOOTER_AFTER_ROW:
            cfg.after_add_row = val
    return cfg


def _read_data(ws: Worksheet, first_data_row: int) -> list[list[Any]]:
    rows = list(ws.iter_rows(values_only=True))
    start = max(first_data_row - 1, 1)  # пропустить заголовки
    out: list[list[Any]] = []
    for r in rows[start:]:
        if r is None:
            continue
        values = list(r)
        if any(v not in (None, '') for v in values):
            while values and values[-1] in (None, ''):
                values.pop()
            out.append(values)
    return out


def _parse_type(value: Any) -> TypeSpec:
    if value is None:
        return TypeSpec(kinds=(KIND_STRING,))
    text = str(value).strip()
    if not text:
        return TypeSpec(kinds=(KIND_STRING,))
    try:
        return parse_type_desc(text)
    except ValueError:
        return TypeSpec(kinds=(KIND_STRING,))


def _flag(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return bool(text) and text != '0'


def _flag_text(flag: bool) -> str:
    return 'Х' if flag else ''


def _int(value: Any) -> int:
    if value is None:
        return 0
    try:
        return int(str(value).strip())
    except ValueError:
        return 0
