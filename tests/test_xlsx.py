"""Unit-тесты xlsx-отчёта."""
from openpyxl import load_workbook

from onec_converter import xlsx_report
from onec_converter.intermediate import make_object


def test_report_creates_sheet(tmp_path):
    objs = [
        make_object('Справочник.Банки', '1|', ['0001', 'Банк'], {'Код': '0001', 'Имя': 'Банк'}, {}),
        make_object('Справочник.Банки', '2|', ['0002', 'Филиал'], {'Код': '0002', 'Имя': 'Филиал'}, {}),
    ]
    p = tmp_path / 'report.xlsx'
    xlsx_report.build_report(objs, p)
    wb = load_workbook(p)
    ws = wb.active
    assert ws['A1'].value == 'Ключ'
    assert ws['A2'].value == '0001|Банк'
    assert ws.max_row == 3
