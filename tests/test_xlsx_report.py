"""Фаза 8: unit-тесты XLSX-отчётов (структура + размеры).

openpyxl читает файл обратно: проверка листов, заголовков, строк,
кириллицы, пустых структур и top_n.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from onec_converter.xlsx_report import build_sizes_report, build_structure_report


def test_structure_report_sheets_and_rows(tmp_path: Path):
    diff = {
        'only_source': ['Справочник.Банки', 'Документ.Оплата'],
        'only_target': ['Справочник.Клиенты'],
        'type_mismatch': [
            {'object': 'Справочник.Товары', 'attr': 'Цена',
             'source_type': 'Numeric', 'target_type': 'String'},
        ],
        'counts': {'only_source': 2, 'only_target': 1, 'mismatch': 1},
    }
    out = build_structure_report(diff, tmp_path / 'structure.xlsx')
    wb = load_workbook(out)
    assert wb.sheetnames == ['Только в источнике', 'Только в приёмнике',
                             'Расхождения типов']
    ws = wb['Только в источнике']
    assert [c.value for c in ws[1]] == ['Объект']
    assert [ws.cell(row=r, column=1).value for r in (2, 3)] == \
        ['Справочник.Банки', 'Документ.Оплата']
    ws = wb['Расхождения типов']
    assert [c.value for c in ws[1]] == ['Объект', 'Поле', 'Тип источника',
                                        'Тип приёмника']
    assert [ws.cell(row=2, column=c).value for c in range(1, 5)] == \
        ['Справочник.Товары', 'Цена', 'Numeric', 'String']


def test_structure_report_empty(tmp_path: Path):
    """Пустая структура: листы с заголовками и 0 строк."""
    diff = {'only_source': [], 'only_target': [],
            'type_mismatch': [], 'counts': {'only_source': 0, 'only_target': 0,
                                            'mismatch': 0}}
    out = build_structure_report(diff, tmp_path / 'empty.xlsx')
    wb = load_workbook(out)
    assert wb.sheetnames == ['Только в источнике', 'Только в приёмнике',
                             'Расхождения типов']
    for sheet in wb.sheetnames:
        assert wb[sheet].max_row == 1  # только заголовок


def test_structure_report_cyrillic(tmp_path: Path):
    """Кириллица в заголовках листов и значениях не искажается."""
    diff = {'only_source': ['Справочник.Номенклатура «Ковры»'],
            'only_target': [], 'type_mismatch': [],
            'counts': {'only_source': 1, 'only_target': 0, 'mismatch': 0}}
    out = build_structure_report(diff, tmp_path / 'cyr.xlsx')
    wb = load_workbook(out)
    # «» и кириллица в имени листа (31 символ, запрещённые заменены)
    assert wb['Только в источнике'].cell(row=2, column=1).value == \
        'Справочник.Номенклатура «Ковры»'


def test_sizes_report_sorted_top_n(tmp_path: Path):
    sizes = [('t1', 10, 100), ('t2', 5, 5000), ('t3', 1, 50), ('t4', 2, 300)]
    out = build_sizes_report(sizes, tmp_path / 'sizes.xlsx', top_n=2)
    wb = load_workbook(out)
    assert wb.sheetnames == ['Таблицы']
    ws = wb['Таблицы']
    assert [c.value for c in ws[1]] == ['Таблица', 'Строки', 'Байты']
    # топ-2 по байтам: t2 (5000), t4 (300)
    assert [ws.cell(row=2, column=1).value, ws.cell(row=3, column=1).value] == \
        ['t2', 't4']
    assert ws.cell(row=2, column=3).value == 5000


def test_sizes_report_empty(tmp_path: Path):
    out = build_sizes_report([], tmp_path / 'sizes_empty.xlsx')
    wb = load_workbook(out)
    assert wb['Таблицы'].max_row == 1
